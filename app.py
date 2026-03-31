"""
Gino's PLUG Option Pricer
Flask application for pricing options packages with PLUG strikes.
All strikes are assumed to be in % terms.
B/O shifts are configured as % of notional by ticker and expiry month.
"""

import os
import json
import math
import socket
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

# Set timeout for slow pricing models
socket.setdefaulttimeout(600)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Default symbol to RIC mapping
SYMBOL_TO_RIC = {
    'SPY': 'SPY.P',
    'QQQ': 'QQQ.O',
    'IWM': 'IWM.P',
    'DIA': 'DIA.P',
    'AAPL': 'AAPL.O',
    'MSFT': 'MSFT.O',
    'GOOGL': 'GOOGL.O',
    'AMZN': 'AMZN.O',
    'NVDA': 'NVDA.O',
    'META': 'META.O',
    'TSLA': 'TSLA.O',
}

SYMBOLS_FILE = 'symbols_data.json'
SHIFTS_FILE = 'bo_shifts_data.json'

def load_symbols():
    if os.path.exists(SYMBOLS_FILE):
        with open(SYMBOLS_FILE, 'r') as f:
            return json.load(f)
    return SYMBOL_TO_RIC.copy()

def save_symbols(symbols):
    with open(SYMBOLS_FILE, 'w') as f:
        json.dump(symbols, f, indent=2)

def get_ric(symbol):
    symbols = load_symbols()
    return symbols.get(symbol.upper())

def get_expiry_months():
    """Generate list of expiry months for next 2 years"""
    months = []
    current = date.today().replace(day=1)
    for i in range(24):
        month_date = current + relativedelta(months=i)
        months.append(month_date.strftime('%Y-%m'))
    return months

def load_shifts():
    if os.path.exists(SHIFTS_FILE):
        with open(SHIFTS_FILE, 'r') as f:
            data = json.load(f)
            data = ensure_shifts_current(data)
            return data
    
    months = get_expiry_months()
    return {
        'months': months,
        'tickers': {}
    }

def ensure_shifts_current(shifts):
    """Ensure shift months are current, adding new months and removing old ones"""
    current_months = get_expiry_months()
    shifts['months'] = current_months
    
    for ticker in shifts.get('tickers', {}):
        ticker_shifts = shifts['tickers'][ticker]
        new_shifts = {}
        for month in current_months:
            if month in ticker_shifts:
                new_shifts[month] = ticker_shifts[month]
            else:
                existing_months = sorted([m for m in ticker_shifts.keys() if m in current_months or m < current_months[-1]])
                if existing_months:
                    last_month = existing_months[-1]
                    new_shifts[month] = ticker_shifts.get(last_month, 0.0)
                else:
                    new_shifts[month] = 0.0
        shifts['tickers'][ticker] = new_shifts
    
    return shifts

def save_shifts(shifts):
    with open(SHIFTS_FILE, 'w') as f:
        json.dump(shifts, f, indent=2)

def interpolate_bo_shift(expiry_date, months, shift_values):
    """Interpolate B/O shift for a given expiry date"""
    if not months or not shift_values:
        return 0.0
    
    expiry_month = expiry_date.strftime('%Y-%m')
    
    for i, month in enumerate(months):
        if month == expiry_month:
            return shift_values[i]
        if month > expiry_month:
            if i == 0:
                return shift_values[0]
            prev_month = datetime.strptime(months[i-1] + '-01', '%Y-%m-%d').date()
            curr_month = datetime.strptime(month + '-01', '%Y-%m-%d').date()
            days_total = (curr_month - prev_month).days
            days_to_expiry = (expiry_date - prev_month).days
            if days_total > 0:
                ratio = days_to_expiry / days_total
                return shift_values[i-1] + ratio * (shift_values[i] - shift_values[i-1])
            return shift_values[i-1]
    
    return shift_values[-1] if shift_values else 0.0

def create_model(symbol, date_val, quantity, call_put, style, strike, ref_price, side, settle_timing):
    """
    Placeholder for the actual pricing model.
    Replace this with your actual model.
    """
    T = max((date_val - datetime.now().date()).days / 365.0, 0.001)
    r = 0.05
    sigma = 0.20
    
    S = ref_price
    K = strike
    
    if K <= 0 or S <= 0:
        return {
            'Spot': ref_price,
            'Fair Value': 0.0,
            'Delta': 0.0,
            'Vega': 0.0,
            'Rho': 0.0,
            'Borrow01': 0.0
        }
    
    d1 = (math.log(S/K) + (r + sigma**2/2)*T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    
    def norm_cdf(x):
        return 0.5 * (1 + math.erf(x / math.sqrt(2)))
    
    def norm_pdf(x):
        return math.exp(-x**2/2) / math.sqrt(2 * math.pi)
    
    if call_put.upper() in ['CALL', 'C']:
        FV = S * norm_cdf(d1) - K * math.exp(-r*T) * norm_cdf(d2)
        delta = norm_cdf(d1)
    else:
        FV = K * math.exp(-r*T) * norm_cdf(-d2) - S * norm_cdf(-d1)
        delta = -norm_cdf(-d1)
    
    # Clamp delta between -1 and 1
    delta = max(-1.0, min(1.0, delta))
    
    vega = S * norm_pdf(d1) * math.sqrt(T) / 100
    rho = K * T * math.exp(-r*T) * (norm_cdf(d2) if call_put.upper() in ['CALL', 'C'] else -norm_cdf(-d2)) / 100
    borrow01 = S * delta * T / 10000
    
    return {
        'Spot': ref_price,
        'Fair Value': round(FV, 2),
        'Delta': round(delta, 2),  # Round to 2 decimals
        'Vega': round(vega, 4),
        'Rho': round(rho, 4),
        'Borrow01': round(borrow01, 4)
    }

def solve_plug_strike(plug_record, other_records_premium, target_notional, target_bo_pct, ref_price,
                      max_iterations=50, tolerance=0.01):
    """Solve for the PLUG strike using binary search."""
    target_bo = abs(target_notional) * (target_bo_pct / 100)
    
    customer_buying_plug = plug_record['customer_buying']
    mult = plug_record['multiplier']
    qty = plug_record['quantity']
    
    if customer_buying_plug:
        target_plug_premium = target_notional - target_bo - other_records_premium
    else:
        target_plug_premium = other_records_premium - (target_notional - target_bo)
    
    target_fv = abs(target_plug_premium) / (mult * qty)
    
    is_call = plug_record['call_put'].upper() == 'CALL'
    
    low_strike = ref_price * 0.01
    high_strike = ref_price * 3.0
    
    best_strike = ref_price
    best_error = float('inf')
    
    for _ in range(max_iterations):
        mid_strike = (low_strike + high_strike) / 2
        
        try:
            model_result = create_model(
                symbol=plug_record['ric'],
                date_val=plug_record['date'],
                quantity=plug_record['quantity'],
                call_put=plug_record['call_put'],
                style=plug_record['style'],
                strike=mid_strike,
                ref_price=plug_record['ref_price'],
                side=plug_record['side'],
                settle_timing='close'
            )
            
            current_fv = model_result['Fair Value']
            error = abs(current_fv - target_fv)
            
            if error < best_error:
                best_error = error
                best_strike = mid_strike
            
            if error < tolerance:
                cap_pct = mid_strike / ref_price
                return mid_strike, cap_pct
            
            if is_call:
                if current_fv > target_fv:
                    low_strike = mid_strike
                else:
                    high_strike = mid_strike
            else:
                if current_fv > target_fv:
                    high_strike = mid_strike
                else:
                    low_strike = mid_strike
        except:
            high_strike = mid_strike
    
    cap_pct = best_strike / ref_price
    return best_strike, cap_pct

def find_column(df, possible_names):
    """Find a column by checking multiple possible names (case-insensitive)"""
    df_cols_lower = {col.lower().strip(): col for col in df.columns}
    for name in possible_names:
        name_lower = name.lower().strip()
        if name_lower in df_cols_lower:
            return df_cols_lower[name_lower]
    return None

def parse_excel_file(filepath):
    """Parse an Excel file with % strikes and PLUG"""
    df = pd.read_excel(filepath, header=None)
    
    header_row = 0
    for i, row in df.iterrows():
        row_str = ' '.join([str(x).lower() for x in row.values if pd.notna(x)])
        if 'symbol' in row_str or 'package' in row_str:
            header_row = i
            break
    
    df = pd.read_excel(filepath, header=header_row)
    df.columns = [str(col).strip() for col in df.columns]
    
    package_col = find_column(df, ['Package #', 'Package', 'Package#', 'Pkg', 'ETF'])
    symbol_col = find_column(df, ['Symbol', 'Ticker', 'Underlying', 'SYMBOL'])
    style_col = find_column(df, ['Style', 'Option Style', 'STYLE'])
    date_col = find_column(df, ['Date', 'Expiry', 'Expiration', 'Exp Date', 'Expiry Date', 'DATE'])
    strike_col = find_column(df, ['Strike Or %', 'Strike', 'Strike Or Pct', 'Strike/Pct', 'STRIKE'])
    strike_dollar_col = find_column(df, ['Strike or (ref Price * %)', 'Strike Dollar', 'Computed Strike'])
    qty_col = find_column(df, ['Qty', 'Quantity', 'QTY', 'QUANTITY', 'QUANITY'])
    cp_col = find_column(df, ['Call or Put', 'Call/Put', 'C/P', 'Option Type', 'CALL OR PUT'])
    side_col = find_column(df, ['Side', 'Direction', 'B/S', 'SIDE'])
    ref_price_col = find_column(df, ['Ref Price', 'Reference Price', 'Ref', 'Spot', 'RefPrice', 'REFERENCE PRICE', 'Reference'])
    notional_col = find_column(df, ['Notional', 'NOTIONAL', 'Premium', 'Target Premium', 'Client Premium'])
    mult_col = find_column(df, ['Mult', 'Multiplier', 'MULT', 'MULTIPLIER'])
    
    if not all([symbol_col, date_col, strike_col, qty_col, cp_col, side_col]):
        missing = []
        if not symbol_col: missing.append('Symbol')
        if not date_col: missing.append('Date')
        if not strike_col: missing.append('Strike')
        if not qty_col: missing.append('Quantity')
        if not cp_col: missing.append('Call/Put')
        if not side_col: missing.append('Side')
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    
    records = []
    ref_prices = {}
    
    for idx, row in df.iterrows():
        symbol = str(row[symbol_col]).strip().upper() if pd.notna(row.get(symbol_col)) else None
        if not symbol or symbol == 'NAN':
            continue
        
        package = None
        if package_col and pd.notna(row.get(package_col)):
            package = str(row[package_col]).strip()
        
        if ref_price_col and pd.notna(row.get(ref_price_col)):
            ref_val = row[ref_price_col]
            if isinstance(ref_val, (int, float)):
                ref_price = float(ref_val)
                key = (package, symbol)
                ref_prices[key] = ref_price
    
    row_number = 0
    last_values = {}
    
    for idx, row in df.iterrows():
        symbol = str(row[symbol_col]).strip().upper() if pd.notna(row.get(symbol_col)) else None
        
        if not symbol or symbol == 'NAN':
            records.append({'is_blank': True, 'row_index': idx})
            continue
        
        row_number += 1
        
        try:
            package = None
            if package_col and pd.notna(row.get(package_col)):
                package = str(row[package_col]).strip()
            else:
                package = symbol
            
            ref_price = ref_prices.get((package, symbol))
            if ref_price is None:
                for key, price in ref_prices.items():
                    if key[1] == symbol:
                        ref_price = price
                        break
            
            date_val = row[date_col] if date_col else None
            if pd.isna(date_val) or date_val is None:
                if package in last_values and 'date' in last_values[package]:
                    date_val = last_values[package]['date']
            else:
                if isinstance(date_val, str):
                    date_val = pd.to_datetime(date_val).date()
                elif isinstance(date_val, datetime):
                    date_val = date_val.date()
                elif hasattr(date_val, 'date'):
                    date_val = date_val.date()
            
            strike_raw = row[strike_col]
            is_plug = False
            strike_pct = None
            strike = None
            
            if isinstance(strike_raw, str) and strike_raw.strip().upper() == 'PLUG':
                is_plug = True
                strike = None
            else:
                if strike_dollar_col and pd.notna(row.get(strike_dollar_col)):
                    try:
                        computed_strike = float(row[strike_dollar_col])
                        if computed_strike > 0:
                            strike = computed_strike
                            if pd.notna(strike_raw):
                                try:
                                    strike_pct = float(strike_raw)
                                except:
                                    pass
                    except:
                        pass
                
                if strike is None and pd.notna(strike_raw) and ref_price:
                    try:
                        strike_pct = float(strike_raw)
                        strike = ref_price * strike_pct
                    except:
                        strike = None
            
            qty = int(row[qty_col]) if pd.notna(row.get(qty_col)) else 1
            
            cp_raw = str(row[cp_col]).strip().upper() if pd.notna(row.get(cp_col)) else 'CALL'
            call_put = 'CALL' if cp_raw in ['CALL', 'C'] else 'PUT'
            
            style_raw = None
            if style_col and pd.notna(row.get(style_col)):
                style_raw = str(row[style_col]).strip().upper()
            elif package in last_values and 'style_raw' in last_values[package]:
                style_raw = last_values[package]['style_raw']
            else:
                style_raw = 'EUROPEAN'
            
            style = 'European' if 'EUROPEAN' in style_raw else 'American'
            settle = 'CASH' if 'CASH' in style_raw else 'PHYSICAL'
            
            side_raw = str(row[side_col]).strip().upper() if pd.notna(row.get(side_col)) else 'BYO'
            customer_buying = side_raw in ['BYO', 'BYC']
            
            notional = None
            if notional_col and pd.notna(row.get(notional_col)):
                try:
                    notional = float(row[notional_col])
                except:
                    notional = None
            
            multiplier = 100
            if mult_col and pd.notna(row.get(mult_col)):
                try:
                    multiplier = int(row[mult_col])
                except:
                    multiplier = 100
            
            if package not in last_values:
                last_values[package] = {}
            if date_val is not None:
                last_values[package]['date'] = date_val
            if style_raw:
                last_values[package]['style_raw'] = style_raw
            
            records.append({
                'is_blank': False,
                'package': package,
                'symbol': symbol,
                'ric': get_ric(symbol),
                'date': date_val,
                'strike': strike,
                'strike_pct': strike_pct,
                'is_plug': is_plug,
                'quantity': qty,
                'call_put': call_put,
                'style': style,
                'settle': settle,
                'side': side_raw,
                'customer_buying': customer_buying,
                'ref_price': ref_price,
                'multiplier': multiplier,
                'notional': notional,
                'row_index': idx,
                'row_number': row_number
            })
        except Exception as e:
            records.append({
                'is_blank': False,
                'package': str(row.get(package_col, 'Unknown')).strip() if package_col and pd.notna(row.get(package_col)) else 'Unknown',
                'error': str(e),
                'row_index': idx,
                'row_number': row_number
            })
    
    return records

def price_records(records, bo_shifts):
    """Price all records, solving for PLUG strikes"""
    results = []
    package_data = {}
    
    packages = {}
    for record in records:
        if record.get('is_blank'):
            results.append({'is_blank': True, 'row_index': record['row_index']})
            continue
        if 'error' in record:
            results.append({**record, 'priced': False})
            continue
        
        pkg = record.get('package', 'Unknown')
        if pkg not in packages:
            packages[pkg] = {'records': [], 'plug_record': None, 'notional': None}
        
        if record.get('is_plug'):
            packages[pkg]['plug_record'] = record
            if record.get('notional'):
                packages[pkg]['notional'] = record['notional']
        else:
            packages[pkg]['records'].append(record)
            if record.get('notional') and packages[pkg]['notional'] is None:
                packages[pkg]['notional'] = record['notional']
    
    shifts = bo_shifts
    months = shifts.get('months', [])
    
    for pkg, pkg_data in packages.items():
        non_plug_records = pkg_data['records']
        plug_record = pkg_data['plug_record']
        target_notional = pkg_data['notional']
        
        non_plug_results = []
        non_plug_premium = 0
        
        for record in non_plug_records:
            if record.get('ric') is None:
                result = {
                    **record,
                    'priced': False,
                    'error': f"Symbol '{record.get('symbol')}' not found in RIC mapping."
                }
                non_plug_results.append(result)
                results.append(result)
                continue
            
            try:
                model_result = create_model(
                    symbol=record['ric'],
                    date_val=record['date'],
                    quantity=record['quantity'],
                    call_put=record['call_put'],
                    style=record['style'],
                    strike=record['strike'],
                    ref_price=record['ref_price'],
                    side=record['side'],
                    settle_timing='close'
                )
                
                mult = record['multiplier']
                qty = record['quantity']
                
                fair_value = model_result['Fair Value']
                premium = fair_value * mult * qty
                premium_signed = premium if record['customer_buying'] else -premium
                non_plug_premium += premium_signed
                
                delta = model_result['Delta']
                delta = max(-1.0, min(1.0, delta))
                delta = round(delta, 2)
                
                delta_sign = 1 if record['customer_buying'] else -1
                shares_to_hedge = delta * mult * qty * delta_sign
                dollar_delta = delta * qty * mult * record['ref_price'] * delta_sign
                
                vega_per_contract = model_result['Vega']
                rho_per_contract = model_result['Rho']
                borrow01_per_contract = model_result['Borrow01']
                
                side_mult = -1 if record['customer_buying'] else 1
                vega = vega_per_contract * mult * qty * side_mult
                rho = rho_per_contract * mult * qty * side_mult
                borrow01 = borrow01_per_contract * mult * qty * side_mult
                
                date_str = record['date'].strftime('%m/%d/%Y') if record.get('date') else ''
                
                result = {
                    **record,
                    'priced': True,
                    'date': date_str,
                    'date_raw': record['date'].isoformat() if record.get('date') else None,
                    'fair_value': fair_value,
                    'premium': round(premium_signed, 2),
                    'delta': delta,
                    'dollar_delta': round(dollar_delta, 2),
                    'shares_to_hedge': round(shares_to_hedge, 2),
                    'vega': round(vega, 2),
                    'rho': round(rho, 2),
                    'borrow01': round(borrow01, 2),
                    'cap': None
                }
                non_plug_results.append(result)
                results.append(result)
                
            except Exception as e:
                result = {**record, 'priced': False, 'error': str(e)}
                non_plug_results.append(result)
                results.append(result)
        
        if plug_record and target_notional:
            if plug_record.get('ric') is None:
                results.append({
                    **plug_record,
                    'priced': False,
                    'error': f"Symbol '{plug_record.get('symbol')}' not found in RIC mapping."
                })
                continue
            
            try:
                symbol = plug_record['symbol']
                expiry_date = plug_record['date']
                
                target_bo_pct = 0.0
                if symbol in shifts.get('tickers', {}):
                    ticker_shifts = shifts['tickers'][symbol]
                    shift_values = [ticker_shifts.get(m, 0.0) for m in months]
                    target_bo_pct = interpolate_bo_shift(expiry_date, months, shift_values)
                
                solved_strike, cap_pct = solve_plug_strike(
                    plug_record=plug_record,
                    other_records_premium=non_plug_premium,
                    target_notional=target_notional,
                    target_bo_pct=target_bo_pct,
                    ref_price=plug_record['ref_price']
                )
                
                model_result = create_model(
                    symbol=plug_record['ric'],
                    date_val=plug_record['date'],
                    quantity=plug_record['quantity'],
                    call_put=plug_record['call_put'],
                    style=plug_record['style'],
                    strike=solved_strike,
                    ref_price=plug_record['ref_price'],
                    side=plug_record['side'],
                    settle_timing='close'
                )
                
                mult = plug_record['multiplier']
                qty = plug_record['quantity']
                
                fair_value = model_result['Fair Value']
                premium = fair_value * mult * qty
                premium_signed = premium if plug_record['customer_buying'] else -premium
                
                delta = model_result['Delta']
                delta = max(-1.0, min(1.0, delta))
                delta = round(delta, 2)
                
                delta_sign = 1 if plug_record['customer_buying'] else -1
                shares_to_hedge = delta * mult * qty * delta_sign
                dollar_delta = delta * qty * mult * plug_record['ref_price'] * delta_sign
                
                vega_per_contract = model_result['Vega']
                rho_per_contract = model_result['Rho']
                borrow01_per_contract = model_result['Borrow01']
                
                side_mult = -1 if plug_record['customer_buying'] else 1
                vega = vega_per_contract * mult * qty * side_mult
                rho = rho_per_contract * mult * qty * side_mult
                borrow01 = borrow01_per_contract * mult * qty * side_mult
                
                date_str = plug_record['date'].strftime('%m/%d/%Y') if plug_record.get('date') else ''
                
                total_premium = non_plug_premium + premium_signed
                actual_bo = target_notional - total_premium
                actual_bo_pct = (actual_bo / abs(target_notional)) * 100 if target_notional else 0
                
                plug_result = {
                    **plug_record,
                    'priced': True,
                    'date': date_str,
                    'date_raw': plug_record['date'].isoformat() if plug_record.get('date') else None,
                    'strike': solved_strike,
                    'fair_value': fair_value,
                    'premium': round(premium_signed, 2),
                    'delta': delta,
                    'dollar_delta': round(dollar_delta, 2),
                    'shares_to_hedge': round(shares_to_hedge, 2),
                    'vega': round(vega, 2),
                    'rho': round(rho, 2),
                    'borrow01': round(borrow01, 2),
                    'cap': round(cap_pct, 4),
                    'notional': target_notional,
                    'target_bo_pct': round(target_bo_pct, 4),
                    'actual_bo': round(actual_bo, 2),
                    'actual_bo_pct': round(actual_bo_pct, 4)
                }
                results.append(plug_result)
                non_plug_results.append(plug_result)
                
            except Exception as e:
                results.append({**plug_record, 'priced': False, 'error': str(e), 'cap': None})
        elif plug_record:
            results.append({
                **plug_record,
                'priced': False,
                'error': 'PLUG record found but no Notional value specified.',
                'cap': None
            })
        
        pkg_results = [r for r in non_plug_results if r.get('priced')]
        if pkg_results:
            package_data[pkg] = {
                'total_vega': sum(r.get('vega', 0) for r in pkg_results),
                'total_rho': sum(r.get('rho', 0) for r in pkg_results),
                'total_dollar_delta': sum(r.get('dollar_delta', 0) for r in pkg_results),
                'total_borrow01': sum(r.get('borrow01', 0) for r in pkg_results),
                'total_premium': sum(r.get('premium', 0) for r in pkg_results),
                'shares_by_symbol': {},
                'rho_by_date': {},
                'vega_by_ticker_expiry': {},
                'notional': target_notional,
                'lines': pkg_results
            }
            
            for r in pkg_results:
                symbol = r.get('symbol')
                date_raw = r.get('date_raw')
                
                if symbol:
                    if symbol not in package_data[pkg]['shares_by_symbol']:
                        package_data[pkg]['shares_by_symbol'][symbol] = 0
                    package_data[pkg]['shares_by_symbol'][symbol] += r.get('shares_to_hedge', 0)
                
                # Rho by date
                if date_raw:
                    if date_raw not in package_data[pkg]['rho_by_date']:
                        package_data[pkg]['rho_by_date'][date_raw] = 0
                    package_data[pkg]['rho_by_date'][date_raw] += r.get('rho', 0)
                
                # Vega by ticker and expiry
                if symbol and date_raw:
                    key = f"{symbol}|{date_raw}"
                    if key not in package_data[pkg]['vega_by_ticker_expiry']:
                        package_data[pkg]['vega_by_ticker_expiry'][key] = 0
                    package_data[pkg]['vega_by_ticker_expiry'][key] += r.get('vega', 0)
    
    # Add package totals and client input prices
    for result in results:
        if result.get('is_blank') or not result.get('priced'):
            continue
        pkg = result.get('package')
        if pkg in package_data:
            result['package_total_vega'] = round(package_data[pkg]['total_vega'], 2)
            result['package_total_rho'] = round(package_data[pkg]['total_rho'], 2)
            result['package_total_dollar_delta'] = round(package_data[pkg]['total_dollar_delta'], 2)
            result['package_total_premium'] = round(package_data[pkg]['total_premium'], 2)
            result['shares_by_symbol'] = {k: round(v, 2) for k, v in package_data[pkg]['shares_by_symbol'].items()}
            result['rho_by_date'] = {k: round(v, 2) for k, v in package_data[pkg]['rho_by_date'].items()}
            result['vega_by_ticker_expiry'] = {k: round(v, 2) for k, v in package_data[pkg]['vega_by_ticker_expiry'].items()}
            
            # Calculate client input price
            notional = package_data[pkg].get('notional')
            total_premium = package_data[pkg]['total_premium']
            
            if notional and total_premium != 0:
                premium = result.get('premium', 0)
                if premium != 0:
                    client_premium = notional * (premium / total_premium)
                    mult = result.get('multiplier', 100)
                    qty = result.get('quantity', 1)
                    sign = 1 if result.get('customer_buying') else -1
                    client_price = client_premium / (mult * qty * sign)
                    result['client_input_price'] = round(client_price, 2)
                else:
                    result['client_input_price'] = 0.0
            else:
                result['client_input_price'] = result.get('fair_value', 0)
    
    return results, package_data

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/pricing')
def pricing():
    today = date.today().strftime('%Y-%m-%d')
    output_files = []
    for f in os.listdir(app.config['OUTPUT_FOLDER']):
        if f.endswith('.json') and today in f:
            output_files.append(f)
    output_files.sort(reverse=True)
    return render_template('pricing.html', output_files=output_files)

@app.route('/bo_shifts')
def bo_shifts_page():
    shifts = load_shifts()
    shifts = ensure_shifts_current(shifts)
    save_shifts(shifts)
    symbols = load_symbols()
    tickers = list(symbols.keys())
    return render_template('bo_shifts.html', shifts=shifts, tickers=tickers, months=shifts['months'])

@app.route('/symbols')
def symbols_page():
    symbols = load_symbols()
    return render_template('symbols.html', symbols=symbols)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        
        records = parse_excel_file(filepath)
        shifts = load_shifts()
        results, package_data = price_records(records, shifts)
        
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_filename = f"{file.filename.rsplit('.', 1)[0]}_{timestamp}.json"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        
        with open(output_path, 'w') as f:
            json.dump({'results': results, 'filename': file.filename}, f, default=str)
        
        return jsonify({
            'success': True,
            'results': results,
            'filename': output_filename
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_output/<filename>')
def get_output(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    with open(filepath, 'r') as f:
        data = json.load(f)
    return jsonify(data)

@app.route('/api/download_excel/<filename>')
def download_excel(filename):
    json_filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(json_filepath):
        return jsonify({'error': 'File not found'}), 404
    
    with open(json_filepath, 'r') as f:
        data = json.load(f)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing Results"
    
    headers = ['Package', 'Symbol', 'Date', 'Strike', 'Strike %', 'Call/Put', 'Qty', 'Side',
               'Fair Value', 'Client Price', 'Delta', '$ Delta', 'Shares to Hedge', 'Vega', 'Rho', 'Borrow01',
               'Premium', 'Cap', 'Notional', 'Target B/O %', 'Actual B/O', 'Actual B/O %',
               'Pkg Total Vega', 'Pkg Total Rho', 'Pkg Total $ Delta', 'Pkg Total Premium']
    
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    currency_format = '#,##0.00'
    number_format = '#,##0.00'
    delta_format = '0.00'
    
    row_num = 2
    for result in data['results']:
        if result.get('is_blank'):
            row_num += 1
            continue
        
        if not result.get('priced'):
            ws.cell(row=row_num, column=1, value=result.get('package', ''))
            ws.cell(row=row_num, column=2, value=result.get('error', 'Pricing failed'))
            row_num += 1
            continue
        
        is_plug = result.get('cap') is not None
        
        ws.cell(row=row_num, column=1, value=result.get('package', ''))
        ws.cell(row=row_num, column=2, value=result.get('symbol', ''))
        ws.cell(row=row_num, column=3, value=result.get('date', ''))
        
        strike_cell = ws.cell(row=row_num, column=4, value=result.get('strike', ''))
        if result.get('strike'):
            strike_cell.number_format = number_format
        
        strike_pct = result.get('strike_pct') or (result.get('cap') if result.get('cap') else None)
        if strike_pct:
            ws.cell(row=row_num, column=5, value=f"{strike_pct*100:.2f}%")
        
        ws.cell(row=row_num, column=6, value=result.get('call_put', ''))
        ws.cell(row=row_num, column=7, value=result.get('quantity', ''))
        ws.cell(row=row_num, column=8, value=result.get('side', ''))
        
        fv_cell = ws.cell(row=row_num, column=9, value=result.get('fair_value', ''))
        if result.get('fair_value') is not None:
            fv_cell.number_format = number_format
        
        cp_cell = ws.cell(row=row_num, column=10, value=result.get('client_input_price', ''))
        if result.get('client_input_price') is not None:
            cp_cell.number_format = number_format
        
        delta_cell = ws.cell(row=row_num, column=11, value=result.get('delta', ''))
        if result.get('delta') is not None:
            delta_cell.number_format = delta_format
        
        dd_cell = ws.cell(row=row_num, column=12, value=result.get('dollar_delta', ''))
        if result.get('dollar_delta') is not None:
            dd_cell.number_format = currency_format
        
        shares_cell = ws.cell(row=row_num, column=13, value=result.get('shares_to_hedge', ''))
        if result.get('shares_to_hedge') is not None:
            shares_cell.number_format = number_format
        
        vega_cell = ws.cell(row=row_num, column=14, value=result.get('vega', ''))
        if result.get('vega') is not None:
            vega_cell.number_format = number_format
        
        rho_cell = ws.cell(row=row_num, column=15, value=result.get('rho', ''))
        if result.get('rho') is not None:
            rho_cell.number_format = number_format
        
        borrow_cell = ws.cell(row=row_num, column=16, value=result.get('borrow01', ''))
        if result.get('borrow01') is not None:
            borrow_cell.number_format = number_format
        
        prem_cell = ws.cell(row=row_num, column=17, value=result.get('premium', ''))
        if result.get('premium') is not None:
            prem_cell.number_format = currency_format
        
        cap_val = result.get('cap')
        if cap_val:
            ws.cell(row=row_num, column=18, value=f"{cap_val*100:.2f}%")
        
        notional_cell = ws.cell(row=row_num, column=19, value=result.get('notional', ''))
        if result.get('notional') is not None:
            notional_cell.number_format = currency_format
        
        if result.get('target_bo_pct') is not None:
            ws.cell(row=row_num, column=20, value=f"{result['target_bo_pct']:.4f}%")
        
        actual_bo_cell = ws.cell(row=row_num, column=21, value=result.get('actual_bo', ''))
        if result.get('actual_bo') is not None:
            actual_bo_cell.number_format = currency_format
        
        if result.get('actual_bo_pct') is not None:
            ws.cell(row=row_num, column=22, value=f"{result['actual_bo_pct']:.4f}%")
        
        ptv_cell = ws.cell(row=row_num, column=23, value=result.get('package_total_vega', ''))
        if result.get('package_total_vega') is not None:
            ptv_cell.number_format = number_format
        
        ptr_cell = ws.cell(row=row_num, column=24, value=result.get('package_total_rho', ''))
        if result.get('package_total_rho') is not None:
            ptr_cell.number_format = number_format
        
        ptdd_cell = ws.cell(row=row_num, column=25, value=result.get('package_total_dollar_delta', ''))
        if result.get('package_total_dollar_delta') is not None:
            ptdd_cell.number_format = currency_format
        
        ptp_cell = ws.cell(row=row_num, column=26, value=result.get('package_total_premium', ''))
        if result.get('package_total_premium') is not None:
            ptp_cell.number_format = currency_format
        
        if is_plug:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = blue_fill
        elif result.get('package_total_premium') and abs(float(result.get('package_total_premium', 0))) > 8000000:
            ws.cell(row=row_num, column=1).fill = yellow_fill
        
        row_num += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    excel_filename = filename.replace('.json', '.xlsx')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=excel_filename)

@app.route('/api/save_bo_shifts', methods=['POST'])
def save_bo_shifts():
    data = request.json
    shifts = load_shifts()
    
    ticker = data.get('ticker')
    shift_values = data.get('shifts', {})
    
    if ticker not in shifts['tickers']:
        shifts['tickers'][ticker] = {}
    
    for month, value in shift_values.items():
        shifts['tickers'][ticker][month] = float(value) if value else 0.0
    
    save_shifts(shifts)
    return jsonify({'success': True})

@app.route('/api/get_bo_shifts')
def get_bo_shifts():
    shifts = load_shifts()
    shifts = ensure_shifts_current(shifts)
    return jsonify(shifts)

@app.route('/api/save_symbols', methods=['POST'])
def save_symbols_api():
    data = request.json
    symbols = data.get('symbols', {})
    save_symbols(symbols)
    return jsonify({'success': True})

@app.route('/api/get_symbols')
def get_symbols():
    symbols = load_symbols()
    return jsonify(symbols)

@app.route('/api/output_files')
def list_output_files():
    today = date.today().strftime('%Y-%m-%d')
    output_files = []
    for f in os.listdir(app.config['OUTPUT_FOLDER']):
        if f.endswith('.json') and today in f:
            output_files.append(f)
    output_files.sort(reverse=True)
    return jsonify(output_files)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=1237)
